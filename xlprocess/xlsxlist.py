import openpyxl


def xlsxtolist(filename='', sheetnum=0):
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[sheetnum]
    datas = []
    datas.append(wb.get_sheet_names()[sheetnum])
    for row in ws.rows:
        ro = []
        for cell in row:
            if cell.value == None:
                ro.append('')
            else:
                ro.append(cell.value)
        datas.append(ro)
    return datas


def xlsxalltolist(filename=''):
    wb = openpyxl.load_workbook(filename)
    sheetnum = len(wb.worksheets)
    datas1 = []
    for a in range(sheetnum):
        ws = wb.worksheets[a]
        datas = []
        datas.append(wb.get_sheet_names()[a])
        for row in ws.rows:
            ro = []
            for cell in row:
                if cell.value == None:
                    ro.append('')
                else:
                    ro.append(cell.value)
            datas.append(ro)
        datas1.append(datas)
    return datas1


def listtoxlsx(filename='', sheets=[[]]):
    wb = openpyxl.Workbook()
    for sheet in sheets:
        table = wb.create_sheet(sheet[0])
        for row in sheet[1:]:
            table.append(row)
    wb.remove(wb.worksheets[0])
    wb.save(filename)
